import openpyxl
import json

result = {}

kpi_file = r"projects\rednote\data\rednote KPI definition_20260323.xlsx"
wb = openpyxl.load_workbook(kpi_file, data_only=True)
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    rows = []
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
        rows.append([str(v) if v is not None else "" for v in row])
    result[f"kpi_{sheet_name}"] = {"headers": rows[0] if rows else [], "data": rows[1:] if len(rows) > 1 else []}

req_file = r"projects\rednote\data\Rednote dashboard V1.0 5-30 requirements.xlsx"
wb2 = openpyxl.load_workbook(req_file, data_only=True)
for sheet_name in wb2.sheetnames:
    ws = wb2[sheet_name]
    rows = []
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
        rows.append([str(v) if v is not None else "" for v in row])
    result[f"req_{sheet_name}"] = {"headers": rows[0] if rows else [], "data": rows[1:] if len(rows) > 1 else []}

with open("tests/excel_requirements.json", "w", encoding="utf-8") as f:
    json.dump(result, f, ensure_ascii=False, indent=2)
print("Saved to tests/excel_requirements.json")
