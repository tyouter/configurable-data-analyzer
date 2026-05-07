import os
import tempfile


def prepare_xlsx(filepath: str) -> tuple[str, bool]:
    has_merges = _check_merges(filepath)
    if not has_merges:
        return filepath, False
    filled = _fill_merged(filepath)
    return filled, True


def _check_merges(filepath: str) -> bool:
    from openpyxl import load_workbook
    try:
        wb = load_workbook(filepath, read_only=False, data_only=True)
        for ws in wb.worksheets:
            if list(ws.merged_cells.ranges):
                wb.close()
                return True
        wb.close()
    except Exception:
        pass
    return False


def _fill_merged(filepath: str) -> str:
    from openpyxl import load_workbook

    wb = load_workbook(filepath, read_only=False, data_only=True)

    for ws in wb.worksheets:
        merge_list = list(ws.merged_cells.ranges)
        if not merge_list:
            continue

        patches = []
        for merged_range in merge_list:
            min_row = merged_range.min_row
            min_col = merged_range.min_col
            val = ws.cell(row=min_row, column=min_col).value
            if val is not None:
                for row in range(merged_range.min_row, merged_range.max_row + 1):
                    for col in range(merged_range.min_col, merged_range.max_col + 1):
                        if row != min_row or col != min_col:
                            patches.append((row, col, val))

        for mr in merge_list:
            ws.unmerge_cells(str(mr))

        for row, col, val in patches:
            ws.cell(row=row, column=col).value = val

    tmp_dir = tempfile.gettempdir()
    base = os.path.splitext(os.path.basename(filepath))[0]
    out_path = os.path.join(tmp_dir, f"chatbi_{base}_filled.xlsx")
    wb.save(out_path)
    wb.close()
    return out_path


def read_excel_filled(filepath: str, sheet_name=None, max_colwidth: int = 80) -> str:
    import pandas as pd

    filled_path, was_filled = prepare_xlsx(filepath)
    try:
        xls = pd.ExcelFile(filled_path)
        all_text = []
        sheets = [sheet_name] if sheet_name else xls.sheet_names
        for sheet in sheets:
            df = pd.read_excel(filled_path, sheet_name=sheet)
            text = df.to_string(index=False, max_colwidth=max_colwidth)
            if text.strip():
                all_text.append(f"=== Sheet: {sheet} ===\n{text}")
        return "\n\n".join(all_text)
    finally:
        if was_filled:
            try:
                os.unlink(filled_path)
            except OSError:
                pass
